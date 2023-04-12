###########################################
# class
###########################################
from openpyxl import load_workbook

class Person(object):
	# Constructor
  	def __init__(self, name, id):
  		self.name = name
  		self.id = id
  	def Display(self):
  		print(self.name, self.id)

class Employee(Person):
	def Print(self):
		print("Employee class called")

class ExcelParser(object):
    count = 0
    def __init__(self,Excelfile):
        self.testfile   = Excelfile
    def count_rows(self):
        rowcount = 0
        try:
            wrkBook = load_workbook(filename = self.testfile)
            if wrkBook is None:
                raise Exception("Cannot open Excel file : %s")
            sheet = wrkBook["tests"]
            rowcount = sheet.max_row
        except Exception as e:
            return(str(e))
        return rowcount
    def count_cols(self):
        colcount = 0
        try:
            wrkBook = load_workbook(filename = self.testfile)
            if wrkBook is None:
                raise Exception("Cannot open Excel file : %s")
            sheet = wrkBook["tests"]
            colcount = sheet.max_column
        except Exception as e:
            return(str(e))
        return colcount
    def test_count(self,rowcount,colcount):
        string1 = ''   
        string2 = "" 
        Test_Count = 0
        Test_Step = 0
        try:
            wrkBook = load_workbook(filename = self.testfile)
            if wrkBook is None:
                raise Exception("Cannot open Excel file : %s")
            sheet = wrkBook["tests"]
            for row in range(1, rowcount, 1):
                for col in range(1, colcount, 1):
                    string = sheet.cell(row = row,column= col).value
                    if string == 'Test Case':
                        for row in range(1, rowcount, 1):
                            for col in range(1, 2, 1):
                                string1 = sheet.cell(row = row + 2,column= col).value
                                string2 = str(string1)
                                if any(c.isalpha() for c in string2):
                                    Test_Step += 1
                                else:
                                    Test_Count += 1
            return Test_Count+1
        except Exception as e:
            return(str(e))

if ( __name__ == "__main__"):
	p1 = Person("Manas", 32)
	p1.Display()
	details = Employee("Ashok", 19)
	details.Display()
	details.Print()
	Row = 0
	Col = 0
	TestCount = 0
	obj = ExcelParser('Zgre.xlsx')
	Row = obj.count_rows()
	Col = obj.count_cols()
	print ("No of Rows in the Test Plan is:",Row)
	print ("No of Columns in the Test Plan is:",Col)
	TestCount = obj.test_count(Row,Col)
	print ("No of Test Cases in the Test Plan is:",TestCount)